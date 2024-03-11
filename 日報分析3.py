import os
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook

def extract_d_column(folder_path):
    output_data = []

    for filename in os.listdir(folder_path):
        if filename.endswith((".xlsx", ".xls")):
            excel_file_path = os.path.join(folder_path, filename)
            try:
                wb = load_workbook(excel_file_path)
                ws = wb.active
                for row in ws.iter_rows():
                    if len(row) > 7:
                        output_data.append(row[7].value)
            except Exception as e:
                messagebox.showwarning("Warning", f"{filename} の読み込み中にエラーが発生しました: {e}")

    return output_data

def save_to_txt(data):
    file_path = filedialog.asksaveasfilename(defaultextension=".txt", filetypes=[("Text files", "*.txt")])
    if file_path:
        try:
            with open(file_path, 'w', encoding='utf-8') as file:
                for item in data:
                    file.write("%s\n" % item)
            messagebox.showinfo("Success", "データが保存されました")
        except Exception as e:
            messagebox.showwarning("Warning", f"ファイルの保存中にエラーが発生しました: {e}")

def browse_folder():
    folder_path = filedialog.askdirectory()
    if folder_path:
        data = extract_d_column(folder_path)
        if data:
            save_to_txt(data)
        else:
            messagebox.showinfo("Info", "XLSXファイルが見つかりませんでした")
    else:
        messagebox.showinfo("Info", "フォルダが選択されていません")

def main():
    root = tk.Tk()
    root.title("日報分析")
    root.geometry("250x80")

    browse_button = tk.Button(root, text="フォルダを選択", command=browse_folder)
    browse_button.pack(pady=20)

    root.mainloop()

if __name__ == "__main__":
    main()
import os
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook

# ExcelファイルからD列のデータを抽出する関数
def extract_d_column(folder_path):
    output_data = []

    # フォルダ内のファイルを走査
    for filename in os.listdir(folder_path):
        # Excelファイル（.xlsxまたは.xls）を処理
        if filename.endswith((".xlsx", ".xls")):
            excel_file_path = os.path.join(folder_path, filename)
            try:
                # Excelファイルを読み込む
                wb = load_workbook(excel_file_path)
                ws = wb.active
                # 各行を走査してh列のデータを抽出
                for index, row in enumerate(ws.iter_rows()):
                    if len(row) > 7:
                        # 1回目のループの場合
                        if index == 0:
                            output_data.append(f"{row[7].value} [{filename}]")
                        else:
                            output_data.append(f"{row[1].value} : {row[7].value}")
            except Exception as e:
                # エラーが発生した場合は警告を表示
                messagebox.showwarning("Warning", f"{filename} の読み込み中にエラーが発生しました: {e}")

    return output_data

# データをテキストファイルに保存する関数
def save_to_txt(data):
    # 保存先のファイルパスを取得
    file_path = filedialog.asksaveasfilename(defaultextension=".txt", filetypes=[("Text files", "*.txt")])
    if file_path:
        try:
            # テキストファイルにデータを書き込む
            with open(file_path, 'w', encoding='utf-8') as file:
                for item in data:
                    file.write("%s\n" % item)
            # 保存成功のメッセージを表示
            messagebox.showinfo("Success", "データが保存されました")
        except Exception as e:
            # 保存中にエラーが発生した場合は警告を表示
            messagebox.showwarning("Warning", f"ファイルの保存中にエラーが発生しました: {e}")

# フォルダを選択する関数
def browse_folder():
    # フォルダ選択ダイアログを表示してフォルダパスを取得
    folder_path = filedialog.askdirectory()
    if folder_path:
        # 選択されたフォルダ内のデータを抽出して保存
        data = extract_d_column(folder_path)
        if data:
            save_to_txt(data)
        else:
            # データが見つからなかった場合は情報メッセージを表示
            messagebox.showinfo("Info", "XLSXファイルが見つかりませんでした")
    else:
        # フォルダが選択されなかった場合は情報メッセージを表示
        messagebox.showinfo("Info", "フォルダが選択されていません")

# メイン関数
def main():
    # Tkinterウィンドウの作成
    root = tk.Tk()
    root.title("日報分析")
    root.geometry("250x80")  # ウィンドウのサイズ設定

    # フォルダ選択ボタンの作成
    browse_button = tk.Button(root, text="フォルダを選択", command=browse_folder)
    browse_button.pack(pady=20)

    # Tkinterのイベントループを実行
    root.mainloop()

if __name__ == "__main__":
    main()
